import os
import re
import zipfile
import shutil
import tempfile
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import filedialog, messagebox
from ttkbootstrap import Style
from ttkbootstrap.widgets import Treeview, Button
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import webbrowser
import logging

# Карта податкових груп
TAX_MAP = {
    "1": "А",
    "2": "Б",
    "3": "В",
    "4": "Г",
    "5": "Д",
    "6": "Е",
    "7": "Ж",
    "8": "З",
}

def log_message(text_widget, message, level="INFO"):
    def open_url(url):
        webbrowser.open_new(url)

    text_widget.config(state="normal")
    if level == "INFO":
        logging.info(message)
    elif level == "WARNING":
        logging.warning(message)
    elif level == "ERROR":
        logging.error(message)

    if "http" in message:
        parts = message.split("http", 1)
        text_widget.insert("end", parts[0])
        url = "http" + parts[1]
        tag_name = f"link_{len(message)}"
        text_widget.insert("end", url + "\n", tag_name)
        text_widget.tag_config(tag_name, foreground="white", underline=True)
        text_widget.tag_bind(tag_name, "<Button-1>", lambda e, link=url: open_url(link))
    else:
        text_widget.insert("end", message + "\n")

    text_widget.config(state="disabled")
    text_widget.see("end")
    text_widget.update()


class SalesParserApp:
    def __init__(self, root):
        self.root = root
        self.root.title("XML Парсер Марія-304Т3 (+Податки)")
        self.style = Style("superhero")
        self.sales_data = []
        self.sales_totals_by_date = {}
        self.temp_dir = None

        self.frame = tk.Frame(self.root)
        self.frame.pack(pady=10)

        self.select_btn = Button(self.frame, text="Обрати ZIP", bootstyle="warning", command=self.select_zip)
        self.select_btn.grid(row=0, column=0, padx=10)

        self.export_btn = Button(self.frame, text="Експорт в Excel", bootstyle="success", command=self.export_to_excel, state="disabled")
        self.export_btn.grid(row=0, column=1, padx=10)

        self.progress_var = tk.DoubleVar()
        self.progress = tk.ttk.Progressbar(self.root, variable=self.progress_var, maximum=100)
        self.progress.pack(fill="x", padx=10, pady=(0, 5))

        self.tree = Treeview(self.root, columns=("date", "time", "check", "name", "amount", "type"), show="headings", height=15)
        self.tree.heading("date", text="Дата")
        self.tree.column("date", width=90, anchor="center")
        self.tree.heading("time", text="Час")
        self.tree.column("time", width=80, anchor="center")
        self.tree.heading("check", text="Номер чека")
        self.tree.column("check", width=100, anchor="center")
        self.tree.heading("name", text="Найменування")
        self.tree.column("name", width=300, anchor="w", stretch=True)
        self.tree.heading("amount", text="Сума (грн)")
        self.tree.column("amount", width=90, anchor="e")
        self.tree.heading("type", text="Тип операції")
        self.tree.column("type", width=100, anchor="center")
        self.tree.pack(padx=10, pady=5, fill="both", expand=True)

        self.log_text = tk.Text(self.root, height=6, state="disabled", bg="black", fg="white", wrap="word")
        self.log_text.pack(fill="both", padx=10, pady=(0, 10))

    def select_zip(self):
        zip_path = filedialog.askopenfilename(filetypes=[("ZIP архів", "*.zip")])
        if not zip_path:
            return

        self.sales_data.clear()
        self.sales_totals_by_date.clear()
        self.tree.delete(*self.tree.get_children())

        self.temp_dir = tempfile.mkdtemp()
        log_message(self.log_text, f"📦 Обрано архів: {zip_path}")

        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(self.temp_dir)
            log_message(self.log_text, f"✅ Розпаковано до: {self.temp_dir}")
        except zipfile.BadZipFile:
            messagebox.showerror("Помилка", "ZIP-файл пошкоджено або не підтримується.")
            return

        files = [f for f in os.listdir(self.temp_dir) if f.endswith(".xml")]
        total = len(files)
        if total == 0:
            log_message(self.log_text, "❌ XML-файли не знайдені у ZIP.")
            return

        log_message(self.log_text, f"🔍 Знайдено {total} XML-файлів. Починаємо обробку...")
        self.progress_var.set(0)
        self.progress.update()

        for idx, filename in enumerate(files, 1):
            file_path = os.path.join(self.temp_dir, filename)
            self.parse_file(file_path)
            self.progress_var.set((idx / total) * 100)
            self.progress.update()
            # log_message(self.log_text, f"🗂 Оброблено: {filename}")

        for row in self.sales_data:
            self.tree.insert("", "end", values=row)

        if self.sales_data:
            self.export_btn.config(state="normal")
            log_message(self.log_text, "✅ Всі файли оброблено.")
        else:
            log_message(self.log_text, "⚠️ Файли прочитано, але не знайдено чеків.")

    def parse_file(self, filepath):
        try:
            with open(filepath, encoding="utf-8") as f:
                content = f.read()

            # Обгортаємо фрагменти DAT у кореневий тег
            dat_blocks = re.findall(r'<DAT.*?</DAT>', content, re.DOTALL)

            for block in dat_blocks:
                fake_xml = f"<root>{block}</root>"
                try:
                    root = ET.fromstring(fake_xml)

                    for c_block in root.findall(".//C"):
                        check_type = c_block.attrib.get("T", "0")
                        is_return = check_type == "1"
                        operation_type = "Повернення" if is_return else "Продаж"

                        e = c_block.find(".//E")
                        if e is None:
                            continue

                        # --- Зчитуємо податкові параметри з E ---
                        e_tax_code = e.attrib.get("TX")        # наприклад "2"
                        e_tax_percent = float(e.attrib.get("TXPR", 0))  # наприклад 7.00

                        # --- 1. Основні дані чека ---
                        ts = e.attrib.get("TS", "")
                        check_no = e.attrib.get("NO", "")
                        
                        # Загальна сума чека по тегу E (вже з урахуванням знижок)
                        check_total_sum_raw = int(e.attrib.get("SM", 0))
                        check_total_sum = abs(check_total_sum_raw) / 100

                        if ts and len(ts) == 14:
                            date = f"{ts[:4]}-{ts[4:6]}-{ts[6:8]}"
                            time_str = f"{ts[8:10]}:{ts[10:12]}:{ts[12:]}"
                        else:
                            date = "Невідомо"
                            time_str = ""

                        # --- 2. Ініціалізуємо структуру для дати ---
                        if date not in self.sales_totals_by_date:
                            self.sales_totals_by_date[date] = {
                                "Продаж": 0,
                                "Повернення": 0,
                                "taxes": {} # {"А": {"turnover": 0, "vat": 0}, ...}
                            }

                        self.sales_totals_by_date[date][operation_type] += check_total_sum

                        # --- 3. РОЗРАХУНОК ПОДАТКІВ (P - D) ---
                        # Створюємо тимчасовий словник для поточного чека: {TaxCode: turnover_cents}
                        current_check_tax_turnover = {}
                        current_check_tax_vat = {} # Якщо треба рахувати ПДВ окремо (але Z-звіт це робить сам, ми рахуємо базу)

                        # А) Додаємо товари (P)
                        for p in c_block.findall(".//P"):
                            tx = p.attrib.get("TX")
                            sm = int(p.attrib.get("SM", 0))
                            current_check_tax_turnover[tx] = current_check_tax_turnover.get(tx, 0) + sm

                        # Б) Віднімаємо знижки (D)
                        # Тег D часто має атрибут TX, що вказує, до якої ставки податку відноситься знижка
                        for d in c_block.findall(".//D"):
                            tx = d.attrib.get("TX")
                            # У знижок SM - це сума знижки
                            sm = int(d.attrib.get("SM", 0))
                            if tx in current_check_tax_turnover:
                                current_check_tax_turnover[tx] -= sm
                            else:
                                # Рідкісний випадок: знижка на групу, якої немає в товарах (глобальна знижка)
                                current_check_tax_turnover[tx] = current_check_tax_turnover.get(tx, 0) - sm

                        # --- 4. Записуємо підсумки чека в загальну таблицю дня ---
                        for tax_code, turnover_cents in current_check_tax_turnover.items():
                            if tax_code in TAX_MAP:
                                tax_name = TAX_MAP[tax_code]
                                turnover_val = abs(turnover_cents) / 100  # Переводимо в гривні
                                
                                # Визначаємо процент для цієї групи
                                tax_percent = 0.0
                                if tax_code == e_tax_code:
                                    tax_percent = e_tax_percent

                                # Формула: оборот × ставка / (100 + ставка)
                                vat_val = 0.0
                                if tax_percent > 0:
                                    vat_val = turnover_val * tax_percent / (100 + tax_percent)



                                taxes_dict = self.sales_totals_by_date[date]["taxes"]
                                if tax_name not in taxes_dict:
                                    taxes_dict[tax_name] = {
                                        "turnover": 0.0,
                                        "vat": 0.0,
                                        "percent": tax_percent
                                    }

                                taxes_dict[tax_name]["turnover"] += turnover_val




                        # --- 5. Збираємо товари для відображення в таблиці (без змін) ---
                        for item in c_block.findall(".//P"):
                            name = item.attrib.get("NM", "Без назви")
                            amount_raw = int(item.attrib.get("SM", 0))
                            amount = abs(amount_raw) / 100
                            type_for_item = "Повернення" if is_return else "Продаж"

                            # СТАНЕТ (правильно):
                            self.sales_data.append((date, time_str, check_no, name, f"{amount:.2f}", type_for_item))

                except ET.ParseError as e:
                    log_message(self.log_text, f"❌ Помилка XML у {os.path.basename(filepath)}: {e}", level="ERROR")

        except Exception as e:
            log_message(self.log_text, f"❌ Помилка читання {os.path.basename(filepath)}: {e}", level="ERROR")

    def export_to_excel(self):
        if not self.sales_data:
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel файли", "*.xlsx")]
        )
        if not save_path:
            return

        df = pd.DataFrame(
            self.sales_data,
            columns=["Дата", "Час", "Номер чека", "Найменування", "Сума (грн)", "Тип операції"]
        )
        df = df.sort_values(by=["Дата", "Час"])

        # --- Загальні підсумки за весь період ---
        grand_total_sales = 0.0
        grand_total_returns = 0.0
        grand_taxes = {}  # { "А": {"turnover": 0, "vat": 0, "percent": 20}, ... }

        output_rows = []

        for date, group in df.groupby("Дата"):
            output_rows.extend(group.values.tolist())

            total_sales = self.sales_totals_by_date.get(date, {}).get("Продаж", 0)
            total_returns = self.sales_totals_by_date.get(date, {}).get("Повернення", 0)
            balance = total_sales - total_returns
            taxes = self.sales_totals_by_date.get(date, {}).get("taxes", {})

            grand_total_sales += total_sales
            grand_total_returns += total_returns

            output_rows.append(["", "", "", "--- ПІДСУМКИ ДНЯ ---", "", ""])
            output_rows.append(["", "", "", "Загальний обіг (Продаж)", f"{total_sales:.2f}", ""])

            # Сортуємо по порядку А–З
            for tax_code in sorted(TAX_MAP.values()):
                tax_data = taxes.get(tax_code)
                if not tax_data:
                    continue

                turnover = tax_data.get("turnover", 0.0)
                percent = tax_data.get("percent", 0.0)

                vat_amount = 0.0
                if percent > 0:
                    vat_amount = turnover * percent / (100 + percent)

                vat_amount = round(vat_amount, 2)


                # Накопичення у загальну таблицю
                if tax_code not in grand_taxes:
                    grand_taxes[tax_code] = {
                        "turnover": 0.0,
                        "vat": 0.0,
                        "percent": percent
                    }
                else:
                    # если процент уже есть, но новый ненулевой — обновляем
                    if percent > 0:
                        grand_taxes[tax_code]["percent"] = percent


                grand_taxes[tax_code]["turnover"] += turnover
                grand_taxes[tax_code]["vat"] += vat_amount

                if turnover != 0:
                    output_rows.append([
                        "", "", "",
                        f"Обіг Група {tax_code} ({percent:.0f}%)",
                        f"{turnover:.2f}",
                        ""
                    ])

                if vat_amount != 0:
                    output_rows.append([
                        "", "", "",
                        f"Податок Група {tax_code} ({percent:.0f}%)",
                        f"{vat_amount:.2f}",
                        ""
                    ])

            output_rows.append(["", "", "", "Повернення", f"{total_returns:.2f}", ""])
            output_rows.append(["", "", "", "ЧИСТИЙ БАЛАНС", f"{balance:.2f}", ""])
            output_rows.append(["", "", "", "", "", ""])

        # --- ФІНАЛЬНА ЗВЕДЕНА ТАБЛИЦЯ ---
        output_rows.append(["", "", "", "============================", "", ""])
        output_rows.append(["", "", "", "ЗВЕДЕНА ТАБЛИЦЯ ЗА ПЕРІОД", "", ""])
        output_rows.append(["", "", "", "============================", "", ""])
        output_rows.append(["", "", "", "ЗАГАЛЬНИЙ ПРОДАЖ (ОБІГ)", f"{grand_total_sales:.2f}", ""])

        for tax_code in sorted(grand_taxes.keys()):
            tax_data = grand_taxes[tax_code]
            turnover = tax_data["turnover"]
            
            percent = tax_data.get("percent", 0.0)
            vat_amount = 0.0
            if percent > 0:
                vat_amount = turnover * percent / (100 + percent)

            vat_amount = round(vat_amount, 2)

            if turnover != 0:
                output_rows.append([
                    "", "", "",
                    f"ЗАГАЛЬНИЙ ОБІГ ГРУПА {tax_code} ({percent:.0f}%)",
                    f"{turnover:.2f}",
                    ""
                ])

            if vat_amount != 0:
                output_rows.append([
                    "", "", "",
                    f"ЗАГАЛЬНИЙ ПОДАТОК ГРУПА {tax_code} ({percent:.0f}%)",
                    f"{vat_amount:.2f}",
                    ""
                ])

        output_rows.append(["", "", "", "ЗАГАЛЬНІ ПОВЕРНЕННЯ", f"{grand_total_returns:.2f}", ""])
        output_rows.append([
            "", "", "",
            "ФІНАЛЬНИЙ БАЛАНС (ЧИСТИЙ)",
            f"{(grand_total_sales - grand_total_returns):.2f}",
            ""
        ])

        export_df = pd.DataFrame(
            output_rows,
            columns=["Дата", "Час", "Номер чека", "Найменування", "Сума (грн)", "Тип операції"]
        )

        try:
            export_df.to_excel(save_path, index=False)
        except PermissionError:
            messagebox.showerror("Помилка", "Не вдалося зберегти файл. Можливо, він відкритий у Excel.")
            return

        # --- Форматування ---
        wb = load_workbook(save_path)
        ws = wb.active

        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            label = str(row[3].value) if row[3].value else ""

            if row[5].value == "Повернення":
                for cell in row:
                    cell.fill = red_fill

            if row[2].value is None and ("Група" in label or "Обіг" in label or "БАЛАНС" in label):
                for cell in row:
                    cell.fill = yellow_fill

            if "ЗАГАЛЬНИЙ" in label or "ЗВЕДЕНА" in label or "ФІНАЛЬНИЙ" in label:
                for cell in row:
                    cell.fill = green_fill

        wb.save(save_path)

        log_message(self.log_text, f"💾 Excel-файл збережено з фінальним звітом: {save_path}")

        if messagebox.askyesno("Готово", "Файл збережено. Відкрити зараз?"):
            try:
                os.startfile(save_path)
            except Exception:
                messagebox.showwarning("Увага", "Не вдалося відкрити файл автоматично.")

        if self.temp_dir and os.path.exists(self.temp_dir):
            if messagebox.askyesno("Очищення", "Очистити тимчасові файли?"):
                try:
                    shutil.rmtree(self.temp_dir)
                    log_message(self.log_text, "🧹 Тимчасову папку очищено.")
                    self.temp_dir = None
                except Exception as e:
                    messagebox.showwarning("Увага", f"Не вдалося видалити тимчасові файли.\n{e}")



if __name__ == "__main__":
    root = tk.Tk()
    app = SalesParserApp(root)
    root.mainloop()