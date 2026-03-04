import os
import re
import zipfile
import shutil
import tempfile
import threading
import queue
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import customtkinter as ctk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import webbrowser

# ─── DPI масштабування (Windows) ─────────────────────────────────────────────
try:
    import ctypes
    ctypes.windll.shcore.SetProcessDpiAwareness(2)
except Exception:
    try:
        ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass

ctk.set_widget_scaling(1.0)
ctk.set_window_scaling(1.0)
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ─── Карта податкових груп ────────────────────────────────────────────────────
TAX_MAP = {
    "1": "А", "2": "Б", "3": "В", "4": "Г",
    "5": "Д", "6": "Е", "7": "Ж", "8": "З",
}

# ─── Кольорова палітра ────────────────────────────────────────────────────────
C = {
    "bg_dark":        "#0F1117",
    "bg_card":        "#1A1D2E",
    "bg_card2":       "#141625",
    "accent_blue":    "#4C9EFF",
    "accent_green":   "#2DD4BF",
    "accent_yellow":  "#F59E0B",
    "accent_red":     "#F87171",
    "accent_purple":  "#A78BFA",
    "text_primary":   "#F1F5F9",
    "text_secondary": "#94A3B8",
    "border":         "#2A2D3E",
    "tv_odd":         "#1E2235",
    "tv_even":        "#181B2A",
    "tv_return":      "#3D1515",
    "tv_summary":     "#192010",
    "tv_grand":       "#0F1F1F",
    "tv_date":        "#0D1530",
}


class SalesParserApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("📊 XML Парсер — Марія-304Т3")
        self.geometry("1280x820")
        self.minsize(1000, 680)
        self.configure(fg_color=C["bg_dark"])

        self.sales_data            = []
        self.sales_totals_by_date  = {}
        self.temp_dir              = None
        self._tax_rate_map         = {}
        self._processing           = False
        self._queue                = queue.Queue()

        self._build_ui()
        self._poll_queue()

    # ══════════════════════════════════════════════════════════════════════════
    #  UI
    # ══════════════════════════════════════════════════════════════════════════
    def _build_ui(self):
        # Header
        hdr = ctk.CTkFrame(self, fg_color=C["bg_card"], corner_radius=0, height=64)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text="  📊  XML Парсер  •  Марія-304Т3",
                     font=ctk.CTkFont(family="Consolas", size=20, weight="bold"),
                     text_color=C["accent_blue"]).pack(side="left", padx=24)
        self.status_label = ctk.CTkLabel(hdr, text="Очікування файлу…",
                                         font=ctk.CTkFont(size=13),
                                         text_color=C["text_secondary"])
        self.status_label.pack(side="right", padx=24)

        # Toolbar
        tb = ctk.CTkFrame(self, fg_color=C["bg_card2"], corner_radius=0, height=56)
        tb.pack(fill="x")
        tb.pack_propagate(False)

        self.btn_open = ctk.CTkButton(tb, text="📂  Відкрити ZIP",
            font=ctk.CTkFont(size=13, weight="bold"), width=170, height=36,
            corner_radius=8, fg_color=C["accent_blue"], hover_color="#3B82F6",
            text_color="#FFF", command=self.select_zip)
        self.btn_open.pack(side="left", padx=(16, 8), pady=10)

        self.btn_export = ctk.CTkButton(tb, text="💾  Експорт Excel",
            font=ctk.CTkFont(size=13, weight="bold"), width=170, height=36,
            corner_radius=8, fg_color=C["accent_green"], hover_color="#14B8A6",
            text_color="#0F172A", command=self.export_to_excel, state="disabled")
        self.btn_export.pack(side="left", padx=8, pady=10)

        self.btn_clear = ctk.CTkButton(tb, text="🗑  Очистити",
            font=ctk.CTkFont(size=13), width=130, height=36, corner_radius=8,
            fg_color="#2A2D3E", hover_color="#374151", text_color=C["text_secondary"],
            command=self.clear_data)
        self.btn_clear.pack(side="left", padx=8, pady=10)

        self.rows_count_lbl = ctk.CTkLabel(tb, text="",
            font=ctk.CTkFont(family="Consolas", size=11), text_color=C["text_secondary"])
        self.rows_count_lbl.pack(side="right", padx=16)

        # Прогрес
        self.progress = ctk.CTkProgressBar(self, height=4, corner_radius=0,
            fg_color=C["bg_card2"], progress_color=C["accent_blue"])
        self.progress.set(0)
        self.progress.pack(fill="x")

        # Основна область
        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True)

        left = ctk.CTkFrame(main, fg_color="transparent")
        left.pack(side="left", fill="both", expand=True)
        self._build_treeview(left)

        right = ctk.CTkFrame(main, fg_color=C["bg_card"], corner_radius=0, width=240)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)
        self._build_stats_panel(right)

        self._build_log(self)

    # ─── TREEVIEW ─────────────────────────────────────────────────────────────
    def _build_treeview(self, parent):
        style = ttk.Style()
        style.theme_use("default")
        style.configure("X.Treeview",
            background=C["tv_odd"], foreground=C["text_primary"],
            fieldbackground=C["tv_odd"], borderwidth=0,
            rowheight=24, font=("Consolas", 10))
        style.configure("X.Treeview.Heading",
            background=C["bg_card2"], foreground=C["accent_blue"],
            font=("Consolas", 10, "bold"), borderwidth=0, relief="flat")
        style.map("X.Treeview",
            background=[("selected", "#2A3A6A")],
            foreground=[("selected", "#FFFFFF")])
        style.map("X.Treeview.Heading",
            background=[("active", C["bg_card"])])
        style.layout("X.Treeview", [("X.Treeview.treearea", {"sticky": "nswe"})])

        self.tree = ttk.Treeview(parent,
            columns=("date","time","check","name","amount","type"),
            show="headings", style="X.Treeview", selectmode="browse")

        for col, text, w, anchor, stretch in [
            ("date",   "Дата",         95,  "center", False),
            ("time",   "Час",          80,  "center", False),
            ("check",  "Чек №",        80,  "center", False),
            ("name",   "Найменування", 400, "w",      True),
            ("amount", "Сума, грн",    100, "e",      False),
            ("type",   "Тип",          90,  "center", False),
        ]:
            self.tree.heading(col, text=text, anchor=anchor)
            self.tree.column(col, width=w, minwidth=60, anchor=anchor, stretch=stretch)

        self.tree.tag_configure("odd",      background=C["tv_odd"],     foreground=C["text_primary"])
        self.tree.tag_configure("even",     background=C["tv_even"],    foreground=C["text_primary"])
        self.tree.tag_configure("return",   background=C["tv_return"],  foreground="#FFAAAA")
        self.tree.tag_configure("summary",  background=C["tv_summary"], foreground=C["accent_yellow"])
        self.tree.tag_configure("grand",    background=C["tv_grand"],   foreground=C["accent_green"])
        self.tree.tag_configure("daterow",  background=C["tv_date"],    foreground=C["accent_blue"])

        sb_y = ctk.CTkScrollbar(parent, command=self.tree.yview)
        sb_y.pack(side="right", fill="y")
        sb_x = ctk.CTkScrollbar(parent, orientation="horizontal", command=self.tree.xview)
        sb_x.pack(side="bottom", fill="x")
        self.tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        self.tree.pack(fill="both", expand=True)

    # ─── ПРАВА ПАНЕЛЬ ─────────────────────────────────────────────────────────
    def _build_stats_panel(self, parent):
        ctk.CTkLabel(parent, text="📈  Статистика",
            font=ctk.CTkFont(family="Consolas", size=13, weight="bold"),
            text_color=C["accent_blue"], anchor="w").pack(fill="x", padx=16, pady=(16,8))
        ctk.CTkFrame(parent, fg_color=C["border"], height=1).pack(fill="x", padx=8)

        outer = ctk.CTkFrame(parent, fg_color="transparent")
        outer.pack(fill="both", expand=True)

        sc = tk.Canvas(outer, bg=C["bg_card"], highlightthickness=0, bd=0)
        sb = tk.Scrollbar(outer, orient="vertical", command=sc.yview)
        sb.pack(side="right", fill="y")
        sc.pack(side="left", fill="both", expand=True)
        sc.configure(yscrollcommand=sb.set)
        self.stats_canvas = sc   # зберігаємо посилання для update в _update_stats

        self.stats_frame = ctk.CTkFrame(sc, fg_color="transparent")
        win = sc.create_window((0, 0), window=self.stats_frame, anchor="nw")
        self.stats_frame.bind("<Configure>", lambda e: sc.configure(scrollregion=sc.bbox("all")))
        sc.bind("<Configure>", lambda e: sc.itemconfig(win, width=e.width))
        sc.bind("<MouseWheel>", lambda e: sc.yview_scroll(int(-1*e.delta/120), "units"))

        self._stat_labels = {}
        for key, label, color in [
            ("total_checks",  "📋 Чеків",        C["text_primary"]),
            ("total_sales",   "💰 Продажі",       C["accent_green"]),
            ("total_returns", "↩️ Повернення",     C["accent_red"]),
            ("net_balance",   "⚖️ Чистий баланс", C["accent_yellow"]),
            ("days_count",    "📅 Днів",           C["text_secondary"]),
        ]:
            row = ctk.CTkFrame(self.stats_frame, fg_color=C["bg_card2"], corner_radius=8)
            row.pack(fill="x", padx=12, pady=4)
            ctk.CTkLabel(row, text=label, font=ctk.CTkFont(size=11),
                         text_color=C["text_secondary"], anchor="w").pack(fill="x", padx=10, pady=(6,0))
            v = ctk.CTkLabel(row, text="—",
                             font=ctk.CTkFont(family="Consolas", size=15, weight="bold"),
                             text_color=color, anchor="w")
            v.pack(fill="x", padx=10, pady=(0,6))
            self._stat_labels[key] = v

        ctk.CTkFrame(self.stats_frame, fg_color=C["border"], height=1).pack(fill="x", padx=8, pady=8)
        ctk.CTkLabel(self.stats_frame, text="🧾 Податки",
            font=ctk.CTkFont(family="Consolas", size=12, weight="bold"),
            text_color=C["accent_purple"], anchor="w").pack(fill="x", padx=12, pady=(0,4))
        self.tax_stats_frame = ctk.CTkFrame(self.stats_frame, fg_color="transparent")
        self.tax_stats_frame.pack(fill="x")

    # ─── ЛОГ ──────────────────────────────────────────────────────────────────
    def _build_log(self, parent):
        log_outer = ctk.CTkFrame(parent, fg_color=C["bg_card2"], corner_radius=0, height=120)
        log_outer.pack(fill="x", side="bottom")
        log_outer.pack_propagate(False)
        ctk.CTkLabel(log_outer, text="  🖥  Журнал подій",
            font=ctk.CTkFont(family="Consolas", size=11, weight="bold"),
            text_color=C["text_secondary"], anchor="w").pack(fill="x", padx=8, pady=(4,0))
        self.log_text = tk.Text(log_outer, height=4, bg=C["bg_dark"], fg=C["text_primary"],
            insertbackground="white", font=("Consolas", 10), wrap="word",
            bd=0, relief="flat", state="disabled")
        self.log_text.pack(fill="both", expand=True, padx=8, pady=(2,6))
        self.log_text.tag_config("OK",    foreground=C["accent_green"])
        self.log_text.tag_config("WARN",  foreground=C["accent_yellow"])
        self.log_text.tag_config("ERROR", foreground=C["accent_red"])
        self.log_text.tag_config("INFO",  foreground=C["text_secondary"])

    # ══════════════════════════════════════════════════════════════════════════
    #  QUEUE POLLING
    # ══════════════════════════════════════════════════════════════════════════
    def _poll_queue(self):
        try:
            while True:
                self._handle_msg(self._queue.get_nowait())
        except queue.Empty:
            pass
        self.after(50, self._poll_queue)

    def _handle_msg(self, msg):
        k = msg.get("kind")
        if k == "log":
            self._log_direct(msg["text"], msg.get("level", "INFO"))
        elif k == "progress":
            self.progress.set(msg["value"])
        elif k == "status":
            self.status_label.configure(text=msg["text"])
        elif k == "done":
            self._on_parse_done()
        elif k == "export_enable":
            self.btn_export.configure(state="normal" if self.sales_data else "disabled")
        elif k == "export_done":
            self.btn_export.configure(state="normal")
            if messagebox.askyesno("Готово", "Файл збережено. Відкрити зараз?"):
                try: os.startfile(msg["path"])
                except Exception: pass
            if self.temp_dir and os.path.exists(self.temp_dir):
                if messagebox.askyesno("Очищення", "Видалити тимчасові файли?"):
                    try:
                        shutil.rmtree(self.temp_dir)
                        self._log_direct("🧹 Тимчасові файли видалено.", "OK")
                        self.temp_dir = None
                    except Exception as e:
                        messagebox.showwarning("Увага", str(e))
        elif k == "error":
            self._processing = False
            self.btn_open.configure(state="normal")
            self.btn_export.configure(state="normal" if self.sales_data else "disabled")
            messagebox.showerror("Помилка", msg["text"])

    # ══════════════════════════════════════════════════════════════════════════
    #  ЛОГУВАННЯ
    # ══════════════════════════════════════════════════════════════════════════
    def _log_direct(self, message, level="INFO"):
        self.log_text.config(state="normal")
        self.log_text.insert("end", message + "\n", level)
        self.log_text.config(state="disabled")
        self.log_text.see("end")
        self.status_label.configure(text=message[:72] + ("…" if len(message) > 72 else ""))

    def log(self, message, level="INFO"):
        """Потокобезпечне логування."""
        self._queue.put({"kind": "log", "text": message, "level": level})

    # ══════════════════════════════════════════════════════════════════════════
    #  ВИБІР ZIP
    # ══════════════════════════════════════════════════════════════════════════
    def select_zip(self):
        if self._processing:
            return
        zip_path = filedialog.askopenfilename(filetypes=[("ZIP архів", "*.zip")])
        if not zip_path:
            return

        self.clear_data(silent=True)
        self._processing = True
        self.btn_open.configure(state="disabled")

        self.temp_dir = tempfile.mkdtemp()
        self._log_direct(f"📦 Архів: {os.path.basename(zip_path)}", "INFO")

        try:
            with zipfile.ZipFile(zip_path, "r") as z:
                z.extractall(self.temp_dir)
        except zipfile.BadZipFile:
            messagebox.showerror("Помилка", "ZIP-файл пошкоджено.")
            self._processing = False
            self.btn_open.configure(state="normal")
            return

        files = sorted(f for f in os.listdir(self.temp_dir) if f.endswith(".xml"))
        if not files:
            self._log_direct("❌ XML-файли не знайдено.", "ERROR")
            self._processing = False
            self.btn_open.configure(state="normal")
            return

        self._log_direct(f"🔍 Знайдено {len(files):,} XML-файлів. Обробка у фоні…", "INFO")
        threading.Thread(target=self._parse_worker, args=(files,), daemon=True).start()

    # ══════════════════════════════════════════════════════════════════════════
    #  ПАРСИНГ (ФОНОВИЙ ПОТІК)
    # ══════════════════════════════════════════════════════════════════════════
    def _parse_worker(self, files):
        total = len(files)
        for idx, fn in enumerate(files, 1):
            self._parse_one(os.path.join(self.temp_dir, fn))
            if idx % 10 == 0 or idx == total:
                self._queue.put({"kind": "progress", "value": idx / total})
                self._queue.put({"kind": "status", "text": f"Обробка… {idx:,}/{total:,}"})

        # Перераховуємо ставки після завершення всіх файлів
        tn2code = {v: k for k, v in TAX_MAP.items()}
        for dd in self.sales_totals_by_date.values():
            for tn, td in dd["taxes"].items():
                code = tn2code.get(tn, "")
                pct  = self._tax_rate_map.get(code, 0.0)
                td["pr"]  = f"{pct:.2f}%"
                td["vat"] = td["turnover"] * pct / (100 + pct) if pct > 0 else 0.0

        self._queue.put({"kind": "done"})

    def _parse_one(self, filepath):
        try:
            with open(filepath, encoding="utf-8") as f:
                content = f.read()

            for block in re.findall(r'<DAT.*?</DAT>', content, re.DOTALL):
                try:
                    root = ET.fromstring(f"<root>{block}</root>")
                    for c in root.findall(".//C"):
                        is_ret = c.attrib.get("T", "0") == "1"
                        op     = "Повернення" if is_ret else "Продаж"

                        e = c.find(".//E")
                        if e is None:
                            continue

                        e_code = e.attrib.get("TX", "")
                        e_pct  = float(e.attrib.get("TXPR", 0))
                        if e_code and e_code not in self._tax_rate_map:
                            self._tax_rate_map[e_code] = e_pct

                        ts  = e.attrib.get("TS", "")
                        no  = e.attrib.get("NO", "")
                        tot = abs(int(e.attrib.get("SM", 0))) / 100

                        if ts and len(ts) == 14:
                            date = f"{ts[:4]}-{ts[4:6]}-{ts[6:8]}"
                            time = f"{ts[8:10]}:{ts[10:12]}:{ts[12:]}"
                        else:
                            date, time = "Невідомо", ""

                        if date not in self.sales_totals_by_date:
                            self.sales_totals_by_date[date] = {
                                "Продаж": 0.0, "Повернення": 0.0, "taxes": {}
                            }
                        self.sales_totals_by_date[date][op] += tot

                        # Обороти по групах
                        trn = {}
                        for p in c.findall(".//P"):
                            tx = p.attrib.get("TX", "")
                            trn[tx] = trn.get(tx, 0) + int(p.attrib.get("SM", 0))
                        for d in c.findall(".//D"):
                            tx = d.attrib.get("TX", "")
                            trn[tx] = trn.get(tx, 0) - int(d.attrib.get("SM", 0))

                        for tx_code, cents in trn.items():
                            if tx_code not in TAX_MAP:
                                continue
                            tx_name = TAX_MAP[tx_code]
                            tv  = abs(cents) / 100
                            pct = self._tax_rate_map.get(tx_code, 0.0)
                            vat = tv * pct / (100 + pct) if pct > 0 else 0.0
                            taxes = self.sales_totals_by_date[date]["taxes"]
                            if tx_name not in taxes:
                                taxes[tx_name] = {"turnover": 0.0, "vat": 0.0, "pr": f"{pct:.2f}%"}
                            s = -1 if is_ret else 1
                            taxes[tx_name]["turnover"] += s * tv
                            taxes[tx_name]["vat"]      += s * vat

                        for item in c.findall(".//P"):
                            nm  = item.attrib.get("NM", "Без назви")
                            amt = abs(int(item.attrib.get("SM", 0))) / 100
                            self.sales_data.append((date, time, no, nm, f"{amt:.2f}", op))

                except ET.ParseError as err:
                    self.log(f"❌ XML error {os.path.basename(filepath)}: {err}", "ERROR")
        except Exception as err:
            self.log(f"❌ Читання файлу: {err}", "ERROR")

    # ══════════════════════════════════════════════════════════════════════════
    #  ПІСЛЯ ПАРСИНГУ
    # ══════════════════════════════════════════════════════════════════════════
    def _on_parse_done(self):
        self._processing = False
        self.btn_open.configure(state="normal")

        if not self.sales_data:
            self._log_direct("⚠️ Чеків не знайдено.", "WARN")
            return

        self._log_direct(f"✅ Парсинг завершено. Позицій: {len(self.sales_data):,}. Рендеринг…", "OK")
        self.update_idletasks()

        self._render_table()
        self._update_stats()

        self.btn_export.configure(state="normal")
        self.progress.set(1.0)
        self.rows_count_lbl.configure(text=f"Позицій: {len(self.sales_data):,}")
        self._log_direct(f"✅ Готово. Позицій: {len(self.sales_data):,}", "OK")

    # ══════════════════════════════════════════════════════════════════════════
    #  РЕНДЕР TREEVIEW
    # ══════════════════════════════════════════════════════════════════════════
    def _render_table(self):
        self.tree.delete(*self.tree.get_children())

        df = pd.DataFrame(self.sales_data,
            columns=["Дата","Час","Номер чека","Найменування","Сума (грн)","Тип операції"]
        ).sort_values(by=["Дата","Час"])

        row_idx = 0
        ins = self.tree.insert  # локальна ссилка — швидше в циклі

        for date, group in df.groupby("Дата"):
            ins("", "end", values=(f"── {date} ──","","","","",""), tags=("daterow",))

            for _, row in group.iterrows():
                if row["Тип операції"] == "Повернення":
                    tag = "return"
                else:
                    tag = "odd" if row_idx % 2 == 0 else "even"
                ins("", "end", values=tuple(row), tags=(tag,))
                row_idx += 1

            totals     = self.sales_totals_by_date.get(date, {})
            t_sales    = totals.get("Продаж", 0)
            t_ret      = totals.get("Повернення", 0)
            taxes      = totals.get("taxes", {})

            ins("", "end", values=("","","","Загальний обіг (Продаж)",f"{t_sales:.2f}",""), tags=("summary",))
            for tc, td in sorted(taxes.items()):
                tv  = td.get("turnover", 0.0)
                vat = td.get("vat", 0.0)
                pr  = td.get("pr", "")
                if tv  != 0: ins("", "end", values=("","","",f"  Обіг Група {tc}  ({pr})",f"{tv:.2f}",""), tags=("summary",))
                if vat != 0: ins("", "end", values=("","","",f"  Податок Група {tc}  ({pr})",f"{vat:.2f}",""), tags=("summary",))
            ins("", "end", values=("","","","Повернення",f"{t_ret:.2f}",""), tags=("summary",))
            ins("", "end", values=("","","","ЧИСТИЙ БАЛАНС",f"{t_sales-t_ret:.2f}",""), tags=("summary",))

        # Зведена таблиця
        g_sales   = sum(v.get("Продаж",0)    for v in self.sales_totals_by_date.values())
        g_returns = sum(v.get("Повернення",0) for v in self.sales_totals_by_date.values())
        g_taxes   = self._calc_grand_taxes()

        ins("", "end", values=("","","","▓▓  ЗВЕДЕНА ТАБЛИЦЯ ЗА ВЕСЬ ПЕРІОД  ▓▓","",""), tags=("daterow",))
        ins("", "end", values=("","","","ЗАГАЛЬНИЙ ПРОДАЖ",f"{g_sales:.2f}",""), tags=("grand",))
        for tc, td in sorted(g_taxes.items()):
            tv  = td.get("turnover",0.0)
            vat = td.get("vat",0.0)
            pr  = td.get("pr","")
            if tv  != 0: ins("", "end", values=("","","",f"  ОБІГ ГРУПА {tc}  ({pr})",f"{tv:.2f}",""), tags=("grand",))
            if vat != 0: ins("", "end", values=("","","",f"  ПОДАТОК ГРУПА {tc}  ({pr})",f"{vat:.2f}",""), tags=("grand",))
        ins("", "end", values=("","","","ЗАГАЛЬНІ ПОВЕРНЕННЯ",f"{g_returns:.2f}",""), tags=("grand",))
        ins("", "end", values=("","","","ФІНАЛЬНИЙ БАЛАНС",f"{g_sales-g_returns:.2f}",""), tags=("grand",))

    # ══════════════════════════════════════════════════════════════════════════
    #  HELPER: ЗВЕДЕНІ ПОДАТКИ
    # ══════════════════════════════════════════════════════════════════════════
    def _calc_grand_taxes(self):
        tn2c = {v: k for k, v in TAX_MAP.items()}
        grand = {}
        for dd in self.sales_totals_by_date.values():
            for tn, td in dd.get("taxes", {}).items():
                pct = self._tax_rate_map.get(tn2c.get(tn,""), 0.0)
                pr  = f"{pct:.2f}%"
                tv  = td.get("turnover", 0.0)
                vat = tv * pct / (100 + pct) if pct > 0 else 0.0
                if tn not in grand:
                    grand[tn] = {"turnover": 0.0, "vat": 0.0, "pr": pr}
                grand[tn]["turnover"] += tv
                grand[tn]["vat"]      += vat
                grand[tn]["pr"]        = pr
        return grand

    # ══════════════════════════════════════════════════════════════════════════
    #  СТАТИСТИКА
    # ══════════════════════════════════════════════════════════════════════════
    def _update_stats(self):
        ts = sum(v.get("Продаж",0)    for v in self.sales_totals_by_date.values())
        tr = sum(v.get("Повернення",0) for v in self.sales_totals_by_date.values())

        self._stat_labels["total_checks"].configure( text=f"{len(set((r[0],r[2]) for r in self.sales_data)):,}")
        self._stat_labels["total_sales"].configure(  text=f"{ts:,.2f} ₴")
        self._stat_labels["total_returns"].configure(text=f"{tr:,.2f} ₴")
        self._stat_labels["net_balance"].configure(  text=f"{ts-tr:,.2f} ₴")
        self._stat_labels["days_count"].configure(   text=f"{len(self.sales_totals_by_date)}")

        # CustomTkinter іноді затримує перемальовку — примусово
        for lbl in self._stat_labels.values():
            lbl.update()

        for w in self.tax_stats_frame.winfo_children():
            w.destroy()

        for tn, td in sorted(self._calc_grand_taxes().items()):
            card = ctk.CTkFrame(self.tax_stats_frame, fg_color=C["bg_card2"], corner_radius=8)
            card.pack(fill="x", padx=12, pady=3)
            ctk.CTkLabel(card, text=f"Група {tn}  ({td.get('pr','')})",
                font=ctk.CTkFont(family="Consolas", size=10, weight="bold"),
                text_color=C["accent_purple"], anchor="w").pack(fill="x", padx=10, pady=(6,0))
            ctk.CTkLabel(card, text=f"Обіг:    {td['turnover']:,.2f} ₴",
                font=ctk.CTkFont(family="Consolas", size=10),
                text_color=C["text_secondary"], anchor="w").pack(fill="x", padx=10)
            ctk.CTkLabel(card, text=f"Податок: {td['vat']:,.2f} ₴",
                font=ctk.CTkFont(family="Consolas", size=11, weight="bold"),
                text_color=C["accent_yellow"], anchor="w").pack(fill="x", padx=10, pady=(0,6))

        # Оновлюємо scrollregion Canvas після додавання нових карток
        self.stats_frame.update_idletasks()
        self.stats_canvas.configure(scrollregion=self.stats_canvas.bbox("all"))

    # ══════════════════════════════════════════════════════════════════════════
    #  ОЧИЩЕННЯ
    # ══════════════════════════════════════════════════════════════════════════
    def clear_data(self, silent=False):
        self.sales_data.clear()
        self.sales_totals_by_date.clear()
        self._tax_rate_map.clear()
        self.tree.delete(*self.tree.get_children())
        for lbl in self._stat_labels.values():
            lbl.configure(text="—")
        for w in self.tax_stats_frame.winfo_children():
            w.destroy()
        self.progress.set(0)
        self.rows_count_lbl.configure(text="")
        self.btn_export.configure(state="disabled")
        if not silent:
            self._log_direct("🗑 Дані очищено.", "INFO")
            self.status_label.configure(text="Очікування файлу…")

    # ══════════════════════════════════════════════════════════════════════════
    #  ЕКСПОРТ EXCEL
    # ══════════════════════════════════════════════════════════════════════════
    def export_to_excel(self):
        if not self.sales_data:
            return
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel файли", "*.xlsx")])
        if not save_path:
            return
        self.btn_export.configure(state="disabled")
        self._log_direct("📊 Формування Excel-файлу…", "INFO")
        threading.Thread(target=self._export_worker, args=(save_path,), daemon=True).start()

    def _export_worker(self, save_path):
        try:
            df = pd.DataFrame(self.sales_data,
                columns=["Дата","Час","Номер чека","Найменування","Сума (грн)","Тип операції"]
            ).sort_values(by=["Дата","Час"])

            g_sales = g_ret = 0.0
            rows = []

            for date, group in df.groupby("Дата"):
                rows.extend(group.values.tolist())
                tot    = self.sales_totals_by_date.get(date, {})
                ts     = tot.get("Продаж", 0)
                tr     = tot.get("Повернення", 0)
                taxes  = tot.get("taxes", {})
                g_sales += ts; g_ret += tr

                rows.append(["","","",f"--- ПІДСУМКИ ДНЯ {date} ---","",""])
                rows.append(["","","","Загальний обіг (Продаж)",f"{ts:.2f}",""])
                for tn, td in sorted(taxes.items()):
                    tv  = td.get("turnover",0.0)
                    vat = td.get("vat",0.0)
                    pr  = td.get("pr","")
                    if tv  != 0: rows.append(["","","",f"Обіг Група {tn} ({pr})",f"{tv:.2f}",""])
                    if vat != 0: rows.append(["","","",f"Податок Група {tn} ({pr})",f"{vat:.2f}",""])
                rows.append(["","","","Повернення",f"{tr:.2f}",""])
                rows.append(["","","","ЧИСТИЙ БАЛАНС",f"{ts-tr:.2f}",""])
                rows.append(["","","","","",""])

            gt = self._calc_grand_taxes()
            rows.append(["","","","ЗВЕДЕНА ТАБЛИЦЯ ЗА ВЕСЬ ПЕРІОД","",""])
            rows.append(["","","","ЗАГАЛЬНИЙ ПРОДАЖ",f"{g_sales:.2f}",""])
            for tn, td in sorted(gt.items()):
                tv  = td.get("turnover",0.0)
                vat = td.get("vat",0.0)
                pr  = td.get("pr","")
                if tv  != 0: rows.append(["","","",f"ЗАГАЛЬНИЙ ОБІГ ГРУПА {tn} ({pr})",f"{tv:.2f}",""])
                if vat != 0: rows.append(["","","",f"ЗАГАЛЬНИЙ ПОДАТОК ГРУПА {tn} ({pr})",f"{vat:.2f}",""])
            rows.append(["","","","ЗАГАЛЬНІ ПОВЕРНЕННЯ",f"{g_ret:.2f}",""])
            rows.append(["","","","ФІНАЛЬНИЙ БАЛАНС",f"{g_sales-g_ret:.2f}",""])

            pd.DataFrame(rows, columns=["Дата","Час","Номер чека","Найменування","Сума (грн)","Тип операції"]
                ).to_excel(save_path, index=False)

            wb = load_workbook(save_path)
            ws = wb.active
            fills = {
                "ret":     PatternFill("solid", fgColor="FFCCCC"),
                "summary": PatternFill("solid", fgColor="FFFACD"),
                "grand":   PatternFill("solid", fgColor="C6EFCE"),
                "hdr":     PatternFill("solid", fgColor="1F3864"),
            }
            for cell in ws[1]:
                cell.fill = fills["hdr"]
                cell.font = Font(bold=True, color="FFFFFF", name="Consolas")
                cell.alignment = Alignment(horizontal="center")

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                lbl = str(row[3].value or "")
                op  = str(row[5].value or "")
                if op == "Повернення":
                    for cell in row: cell.fill = fills["ret"]
                elif any(k in lbl for k in ("ПІДСУМКИ","Обіг","БАЛАНС","Податок","Повернення")):
                    for cell in row:
                        cell.fill = fills["summary"]
                        cell.font = Font(bold="БАЛАНС" in lbl, name="Consolas", size=10)
                elif any(k in lbl for k in ("ЗАГАЛЬНИЙ","ЗВЕДЕНА","ФІНАЛЬНИЙ")):
                    for cell in row:
                        cell.fill = fills["grand"]
                        cell.font = Font(bold=True, name="Consolas", size=10)

            for i, w in enumerate([12,10,12,50,14,14], 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            wb.save(save_path)

            self._queue.put({"kind": "log", "text": f"💾 Збережено: {save_path}", "level": "OK"})
            self._queue.put({"kind": "export_done", "path": save_path})

        except PermissionError:
            self._queue.put({"kind": "error", "text": "Файл відкритий у Excel. Закрийте і спробуйте."})
        except Exception as e:
            self._queue.put({"kind": "error", "text": f"Помилка експорту: {e}"})
        finally:
            self._queue.put({"kind": "export_enable"})


if __name__ == "__main__":
    app = SalesParserApp()
    app.mainloop()
