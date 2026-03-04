import os
import re
import zipfile
import shutil
import tempfile
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import webbrowser
import logging

# ─── DPI масштабування (Windows) ─────────────────────────────────────────────
# Вмикає чіткий текст на екранах з масштабом 125%, 150%, 200% (4K)
try:
    import ctypes
    ctypes.windll.shcore.SetProcessDpiAwareness(2)  # Per-Monitor DPI v2
except Exception:
    try:
        ctypes.windll.user32.SetProcessDPIAware()   # Fallback Windows 7/8
    except Exception:
        pass  # Linux / macOS — масштабом керує система

# CustomTkinter автоматично підлаштовується під системний масштаб екрана.
# Значення 1.0 означає "використати системний DPI без додаткового множника".
ctk.set_widget_scaling(1.0)
ctk.set_window_scaling(1.0)

# ─── Налаштування теми ───────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ─── Карта податкових груп ────────────────────────────────────────────────────
TAX_MAP = {
    "1": "А", "2": "Б", "3": "В", "4": "Г",
    "5": "Д", "6": "Е", "7": "Ж", "8": "З",
}

# ─── Кольорова палітра ────────────────────────────────────────────────────────
COLORS = {
    "bg_dark":       "#0F1117",
    "bg_card":       "#1A1D2E",
    "bg_card2":      "#141625",
    "accent_blue":   "#4C9EFF",
    "accent_green":  "#2DD4BF",
    "accent_yellow": "#F59E0B",
    "accent_red":    "#F87171",
    "accent_purple": "#A78BFA",
    "text_primary":  "#F1F5F9",
    "text_secondary":"#94A3B8",
    "border":        "#2A2D3E",
    "row_odd":       "#1E2235",
    "row_even":      "#181B2A",
    "row_return":    "#2D1A1A",
    "row_summary":   "#1A2020",
}


class SalesParserApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("📊 XML Парсер — Марія-304Т3")
        self.geometry("1280x820")
        self.minsize(1000, 680)
        self.configure(fg_color=COLORS["bg_dark"])

        self.sales_data = []
        self.sales_totals_by_date = {}
        self.temp_dir = None
        self._tax_rate_map = {}   # {tax_code_str: percent_float}

        self._build_ui()

    # ══════════════════════════════════════════════════════════════════════════
    #  БУДОВА ІНТЕРФЕЙСУ
    # ══════════════════════════════════════════════════════════════════════════
    def _build_ui(self):
        # ── HEADER ────────────────────────────────────────────────────────────
        header = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], corner_radius=0, height=64)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)

        ctk.CTkLabel(
            header,
            text="  📊  XML Парсер  •  Марія-304Т3",
            font=ctk.CTkFont(family="Consolas", size=20, weight="bold"),
            text_color=COLORS["accent_blue"],
        ).pack(side="left", padx=24, pady=12)

        self.status_label = ctk.CTkLabel(
            header,
            text="Очікування файлу…",
            font=ctk.CTkFont(size=13),
            text_color=COLORS["text_secondary"],
        )
        self.status_label.pack(side="right", padx=24)

        # ── TOOLBAR ───────────────────────────────────────────────────────────
        toolbar = ctk.CTkFrame(self, fg_color=COLORS["bg_card2"], corner_radius=0, height=56)
        toolbar.pack(fill="x")
        toolbar.pack_propagate(False)

        self.btn_open = ctk.CTkButton(
            toolbar,
            text="📂  Відкрити ZIP",
            font=ctk.CTkFont(size=13, weight="bold"),
            width=170, height=36,
            corner_radius=8,
            fg_color=COLORS["accent_blue"],
            hover_color="#3B82F6",
            text_color="#FFFFFF",
            command=self.select_zip,
        )
        self.btn_open.pack(side="left", padx=(16, 8), pady=10)

        self.btn_export = ctk.CTkButton(
            toolbar,
            text="💾  Експорт Excel",
            font=ctk.CTkFont(size=13, weight="bold"),
            width=170, height=36,
            corner_radius=8,
            fg_color=COLORS["accent_green"],
            hover_color="#14B8A6",
            text_color="#0F172A",
            command=self.export_to_excel,
            state="disabled",
        )
        self.btn_export.pack(side="left", padx=8, pady=10)

        self.btn_clear = ctk.CTkButton(
            toolbar,
            text="🗑  Очистити",
            font=ctk.CTkFont(size=13),
            width=130, height=36,
            corner_radius=8,
            fg_color="#2A2D3E",
            hover_color="#374151",
            text_color=COLORS["text_secondary"],
            command=self.clear_data,
        )
        self.btn_clear.pack(side="left", padx=8, pady=10)

        # ── ПРОГРЕС-БАР ───────────────────────────────────────────────────────
        self.progress = ctk.CTkProgressBar(
            self,
            height=4,
            corner_radius=0,
            fg_color=COLORS["bg_card2"],
            progress_color=COLORS["accent_blue"],
        )
        self.progress.set(0)
        self.progress.pack(fill="x")

        # ── ГОЛОВНА ОБЛАСТЬ (таблиця + бокова панель) ─────────────────────────
        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=0, pady=0)

        # Ліва: таблиця
        left_panel = ctk.CTkFrame(main, fg_color="transparent")
        left_panel.pack(side="left", fill="both", expand=True)

        self._build_table(left_panel)

        # Права: статистика
        right_panel = ctk.CTkFrame(
            main,
            fg_color=COLORS["bg_card"],
            corner_radius=0,
            width=240,
        )
        right_panel.pack(side="right", fill="y")
        right_panel.pack_propagate(False)
        self._build_stats_panel(right_panel)

        # ── ЛОГ-ПАНЕЛЬ ────────────────────────────────────────────────────────
        self._build_log(self)

    # ─── ТАБЛИЦЯ ──────────────────────────────────────────────────────────────
    def _build_table(self, parent):
        # Заголовок
        cols_frame = ctk.CTkFrame(parent, fg_color=COLORS["bg_card2"], corner_radius=0, height=36)
        cols_frame.pack(fill="x")
        cols_frame.pack_propagate(False)

        headers = [
            ("Дата",           90,  "center"),
            ("Час",            80,  "center"),
            ("Чек №",          80,  "center"),
            ("Найменування",   370, "w"),
            ("Сума, грн",      100, "e"),
            ("Тип",            90,  "center"),
        ]
        for text, w, anchor in headers:
            lbl = ctk.CTkLabel(
                cols_frame,
                text=text,
                font=ctk.CTkFont(family="Consolas", size=11, weight="bold"),
                text_color=COLORS["accent_blue"],
                width=w,
                anchor=anchor,
            )
            lbl.pack(side="left", padx=(8 if text == "Дата" else 2, 2))

        # Фрейм зі скролом
        table_frame = ctk.CTkFrame(parent, fg_color="transparent", corner_radius=0)
        table_frame.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(
            table_frame,
            bg=COLORS["bg_dark"],
            highlightthickness=0,
            bd=0,
        )
        scrollbar_y = ctk.CTkScrollbar(table_frame, command=self.canvas.yview)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x = ctk.CTkScrollbar(table_frame, orientation="horizontal", command=self.canvas.xview)
        scrollbar_x.pack(side="bottom", fill="x")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        self.rows_frame = ctk.CTkFrame(self.canvas, fg_color="transparent")
        self.canvas_window = self.canvas.create_window((0, 0), window=self.rows_frame, anchor="nw")

        self.rows_frame.bind("<Configure>", self._on_frame_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

    def _on_frame_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas_window, width=event.width)

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    # ─── ПРАВА ПАНЕЛЬ (СТАТИСТИКА) ────────────────────────────────────────────
    def _build_stats_panel(self, parent):
        ctk.CTkLabel(
            parent,
            text="📈  Статистика",
            font=ctk.CTkFont(family="Consolas", size=13, weight="bold"),
            text_color=COLORS["accent_blue"],
            anchor="w",
        ).pack(fill="x", padx=16, pady=(16, 8))

        sep = ctk.CTkFrame(parent, fg_color=COLORS["border"], height=1)
        sep.pack(fill="x", padx=8)

        # Canvas + scrollbar замість CTkScrollableFrame (сумісність зі старими версіями)
        stats_canvas_frame = ctk.CTkFrame(parent, fg_color="transparent")
        stats_canvas_frame.pack(fill="both", expand=True, padx=0, pady=4)

        stats_canvas = tk.Canvas(
            stats_canvas_frame,
            bg=COLORS["bg_card"],
            highlightthickness=0,
            bd=0,
        )
        stats_sb = tk.Scrollbar(stats_canvas_frame, orient="vertical", command=stats_canvas.yview)
        stats_sb.pack(side="right", fill="y")
        stats_canvas.pack(side="left", fill="both", expand=True)
        stats_canvas.configure(yscrollcommand=stats_sb.set)

        self.stats_frame = ctk.CTkFrame(stats_canvas, fg_color="transparent")
        stats_win = stats_canvas.create_window((0, 0), window=self.stats_frame, anchor="nw")

        def _stats_configure(event):
            stats_canvas.configure(scrollregion=stats_canvas.bbox("all"))
        def _stats_canvas_resize(event):
            stats_canvas.itemconfig(stats_win, width=event.width)
        def _stats_mousewheel(event):
            stats_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        self.stats_frame.bind("<Configure>", _stats_configure)
        stats_canvas.bind("<Configure>", _stats_canvas_resize)
        stats_canvas.bind("<MouseWheel>", _stats_mousewheel)

        self._stat_labels = {}
        fields = [
            ("total_checks",  "📋 Чеків",          COLORS["text_primary"]),
            ("total_sales",   "💰 Продажі",         COLORS["accent_green"]),
            ("total_returns", "↩️ Повернення",       COLORS["accent_red"]),
            ("net_balance",   "⚖️ Чистий баланс",   COLORS["accent_yellow"]),
            ("days_count",    "📅 Днів",             COLORS["text_secondary"]),
        ]
        for key, label_text, color in fields:
            row = ctk.CTkFrame(self.stats_frame, fg_color=COLORS["bg_card2"], corner_radius=8)
            row.pack(fill="x", padx=12, pady=4)

            ctk.CTkLabel(
                row, text=label_text,
                font=ctk.CTkFont(size=11),
                text_color=COLORS["text_secondary"],
                anchor="w",
            ).pack(fill="x", padx=10, pady=(6, 0))

            val_lbl = ctk.CTkLabel(
                row, text="—",
                font=ctk.CTkFont(family="Consolas", size=15, weight="bold"),
                text_color=color,
                anchor="w",
            )
            val_lbl.pack(fill="x", padx=10, pady=(0, 6))
            self._stat_labels[key] = val_lbl

        # Розділювач
        ctk.CTkFrame(self.stats_frame, fg_color=COLORS["border"], height=1).pack(
            fill="x", padx=8, pady=8
        )

        # Блок для податків (наповнюється динамічно)
        ctk.CTkLabel(
            self.stats_frame,
            text="🧾 Податки",
            font=ctk.CTkFont(family="Consolas", size=12, weight="bold"),
            text_color=COLORS["accent_purple"],
            anchor="w",
        ).pack(fill="x", padx=12, pady=(0, 4))

        self.tax_stats_frame = ctk.CTkFrame(self.stats_frame, fg_color="transparent")
        self.tax_stats_frame.pack(fill="x")

    # ─── ЛОГ ──────────────────────────────────────────────────────────────────
    def _build_log(self, parent):
        log_outer = ctk.CTkFrame(parent, fg_color=COLORS["bg_card2"], corner_radius=0, height=130)
        log_outer.pack(fill="x", side="bottom")
        log_outer.pack_propagate(False)

        ctk.CTkLabel(
            log_outer,
            text="  🖥  Журнал подій",
            font=ctk.CTkFont(family="Consolas", size=11, weight="bold"),
            text_color=COLORS["text_secondary"],
            anchor="w",
        ).pack(fill="x", padx=8, pady=(4, 0))

        self.log_text = tk.Text(
            log_outer,
            height=5,
            bg=COLORS["bg_dark"],
            fg=COLORS["text_primary"],
            insertbackground="white",
            font=("Consolas", 10),
            wrap="word",
            bd=0,
            relief="flat",
            state="disabled",
        )
        self.log_text.pack(fill="both", expand=True, padx=8, pady=(2, 6))

        self.log_text.tag_config("INFO",    foreground=COLORS["text_secondary"])
        self.log_text.tag_config("OK",      foreground=COLORS["accent_green"])
        self.log_text.tag_config("WARN",    foreground=COLORS["accent_yellow"])
        self.log_text.tag_config("ERROR",   foreground=COLORS["accent_red"])
        self.log_text.tag_config("LINK",    foreground=COLORS["accent_blue"], underline=True)

    # ══════════════════════════════════════════════════════════════════════════
    #  ЛОГУВАННЯ
    # ══════════════════════════════════════════════════════════════════════════
    def log(self, message, level="INFO"):
        self.log_text.config(state="normal")
        tag = level
        if "http" in message:
            parts = message.split("http", 1)
            self.log_text.insert("end", parts[0], tag)
            url = "http" + parts[1]
            link_tag = f"link_{id(url)}"
            self.log_text.tag_config(link_tag, foreground=COLORS["accent_blue"], underline=True)
            self.log_text.tag_bind(link_tag, "<Button-1>", lambda e, u=url: webbrowser.open_new(u))
            self.log_text.insert("end", url + "\n", link_tag)
        else:
            self.log_text.insert("end", message + "\n", tag)
        self.log_text.config(state="disabled")
        self.log_text.see("end")
        self.log_text.update()
        self.status_label.configure(text=message[:70] + ("…" if len(message) > 70 else ""))

    # ══════════════════════════════════════════════════════════════════════════
    #  РЯДКИ ТАБЛИЦІ
    # ══════════════════════════════════════════════════════════════════════════
    def _add_row(self, values, row_type="normal"):
        idx = len(self.rows_frame.winfo_children())
        if row_type == "return":
            bg = COLORS["row_return"]
        elif row_type == "summary":
            bg = COLORS["row_summary"]
        elif row_type == "grand":
            bg = "#162020"
        else:
            bg = COLORS["row_odd"] if idx % 2 == 0 else COLORS["row_even"]

        row_frame = ctk.CTkFrame(self.rows_frame, fg_color=bg, corner_radius=0, height=26)
        row_frame.pack(fill="x")
        row_frame.pack_propagate(False)

        widths = [90, 80, 80, 370, 100, 90]
        anchors = ["center", "center", "center", "w", "e", "center"]
        text_colors_map = {
            "return":  [COLORS["text_secondary"]] * 5 + [COLORS["accent_red"]],
            "summary": [COLORS["text_secondary"]] * 3 + [COLORS["accent_yellow"], COLORS["accent_green"], COLORS["text_secondary"]],
            "grand":   ["transparent"] * 3 + [COLORS["accent_purple"], COLORS["accent_green"], "transparent"],
            "normal":  [COLORS["text_secondary"], COLORS["text_secondary"], COLORS["text_secondary"],
                        COLORS["text_primary"], COLORS["accent_green"], COLORS["text_secondary"]],
        }
        tc = text_colors_map.get(row_type, text_colors_map["normal"])

        for i, (val, w, anc) in enumerate(zip(values, widths, anchors)):
            c = tc[i] if tc[i] != "transparent" else COLORS["text_secondary"]
            lbl = ctk.CTkLabel(
                row_frame,
                text=str(val) if val is not None else "",
                font=ctk.CTkFont(family="Consolas", size=11, weight="bold" if row_type in ("summary", "grand") else "normal"),
                text_color=c,
                width=w,
                anchor=anc,
            )
            lbl.pack(side="left", padx=(8 if i == 0 else 2, 2))

        return row_frame

    def _add_separator(self, color=None):
        ctk.CTkFrame(
            self.rows_frame,
            fg_color=color or COLORS["border"],
            height=1,
            corner_radius=0,
        ).pack(fill="x", pady=1)

    def _add_section_header(self, text, color=None):
        fr = ctk.CTkFrame(self.rows_frame, fg_color=COLORS["bg_card"], corner_radius=4, height=28)
        fr.pack(fill="x", padx=4, pady=2)
        fr.pack_propagate(False)
        ctk.CTkLabel(
            fr, text=f"  {text}",
            font=ctk.CTkFont(family="Consolas", size=11, weight="bold"),
            text_color=color or COLORS["accent_yellow"],
            anchor="w",
        ).pack(fill="both", expand=True, padx=6)

    # ══════════════════════════════════════════════════════════════════════════
    #  ВИБІР ZIP
    # ══════════════════════════════════════════════════════════════════════════
    def select_zip(self):
        zip_path = filedialog.askopenfilename(filetypes=[("ZIP архів", "*.zip")])
        if not zip_path:
            return

        self.clear_data(silent=True)

        self.temp_dir = tempfile.mkdtemp()
        self.log(f"📦 Архів: {os.path.basename(zip_path)}", "INFO")

        try:
            with zipfile.ZipFile(zip_path, 'r') as z:
                z.extractall(self.temp_dir)
        except zipfile.BadZipFile:
            messagebox.showerror("Помилка", "ZIP-файл пошкоджено.")
            return

        files = [f for f in os.listdir(self.temp_dir) if f.endswith(".xml")]
        total = len(files)
        if total == 0:
            self.log("❌ XML-файли не знайдено.", "ERROR")
            return

        self.log(f"🔍 Знайдено {total} XML-файлів. Обробка…", "INFO")
        self.progress.set(0)

        for idx, filename in enumerate(sorted(files), 1):
            self.parse_file(os.path.join(self.temp_dir, filename))
            self.progress.set(idx / total)
            self.progress.update()

        # ── Після парсингу всіх файлів: оновлюємо pr у всіх записах ──────────
        # TAX_MAP обернений: {"А":"1", "Б":"2", ...} для зворотного пошуку
        tax_name_to_code = {v: k for k, v in TAX_MAP.items()}
        for date_data in self.sales_totals_by_date.values():
            for tax_name, td in date_data["taxes"].items():
                code = tax_name_to_code.get(tax_name, "")
                pct  = self._tax_rate_map.get(code, 0.0)
                td["pr"] = f"{pct:.2f}%"
                # Перераховуємо VAT з правильною ставкою
                tv = td["turnover"]
                td["vat"] = tv * pct / (100 + pct) if pct > 0 else 0.0

        self._render_table()
        self._update_stats()

        if self.sales_data:
            self.btn_export.configure(state="normal")
            self.log(f"✅ Готово. Чеків: {len(self.sales_data)}", "OK")
        else:
            self.log("⚠️ Чеків не знайдено.", "WARN")

    # ══════════════════════════════════════════════════════════════════════════
    #  ПАРСИНГ XML
    # ══════════════════════════════════════════════════════════════════════════
    def parse_file(self, filepath):
        try:
            with open(filepath, encoding="utf-8") as f:
                content = f.read()

            dat_blocks = re.findall(r'<DAT.*?</DAT>', content, re.DOTALL)

            for block in dat_blocks:
                try:
                    root = ET.fromstring(f"<root>{block}</root>")

                    for c_block in root.findall(".//C"):
                        check_type   = c_block.attrib.get("T", "0")
                        is_return    = check_type == "1"
                        operation    = "Повернення" if is_return else "Продаж"

                        e = c_block.find(".//E")
                        if e is None:
                            continue

                        # ── Податкова ставка ──────────────────────────────────
                        e_tax_code    = e.attrib.get("TX", "")           # "2"
                        e_tax_percent = float(e.attrib.get("TXPR", 0))   # 7.00

                        # Зберігаємо в глобальну карту ставок
                        if e_tax_code and e_tax_code not in self._tax_rate_map:
                            self._tax_rate_map[e_tax_code] = e_tax_percent

                        ts       = e.attrib.get("TS", "")
                        check_no = e.attrib.get("NO", "")

                        check_total_raw = int(e.attrib.get("SM", 0))
                        check_total     = abs(check_total_raw) / 100

                        if ts and len(ts) == 14:
                            date     = f"{ts[:4]}-{ts[4:6]}-{ts[6:8]}"
                            time_str = f"{ts[8:10]}:{ts[10:12]}:{ts[12:]}"
                        else:
                            date     = "Невідомо"
                            time_str = ""

                        if date not in self.sales_totals_by_date:
                            self.sales_totals_by_date[date] = {
                                "Продаж": 0.0, "Повернення": 0.0, "taxes": {}
                            }

                        self.sales_totals_by_date[date][operation] += check_total

                        # ── Розрахунок оборотів по групах ────────────────────
                        current_turnover = {}

                        for p in c_block.findall(".//P"):
                            tx = p.attrib.get("TX", "")
                            sm = int(p.attrib.get("SM", 0))
                            current_turnover[tx] = current_turnover.get(tx, 0) + sm

                        for d in c_block.findall(".//D"):
                            tx = d.attrib.get("TX", "")
                            sm = int(d.attrib.get("SM", 0))
                            if tx in current_turnover:
                                current_turnover[tx] -= sm
                            else:
                                current_turnover[tx] = current_turnover.get(tx, 0) - sm

                        # ── Запис у підсумки дня ──────────────────────────────
                        for tax_code, turnover_cents in current_turnover.items():
                            if tax_code not in TAX_MAP:
                                continue
                            tax_name   = TAX_MAP[tax_code]
                            turnover_v = abs(turnover_cents) / 100

                            # Беремо ставку з глобальної карти (вже заповнена вище)
                            tax_pct = self._tax_rate_map.get(tax_code, 0.0)
                            pr_str  = f"{tax_pct:.2f}%"

                            vat = turnover_v * tax_pct / (100 + tax_pct) if tax_pct > 0 else 0.0

                            taxes = self.sales_totals_by_date[date]["taxes"]
                            if tax_name not in taxes:
                                taxes[tax_name] = {
                                    "turnover": 0.0,
                                    "vat":      0.0,
                                    "pr":       pr_str,
                                }
                            else:
                                # Оновлюємо ставку якщо раніше була 0
                                if taxes[tax_name]["pr"] == "0.00%":
                                    taxes[tax_name]["pr"] = pr_str

                            sign = -1 if is_return else 1
                            taxes[tax_name]["turnover"] += sign * turnover_v
                            taxes[tax_name]["vat"]      += sign * vat

                        # ── Зберігаємо позиції для таблиці ───────────────────
                        for item in c_block.findall(".//P"):
                            name   = item.attrib.get("NM", "Без назви")
                            amount = abs(int(item.attrib.get("SM", 0))) / 100
                            self.sales_data.append(
                                (date, time_str, check_no, name, f"{amount:.2f}", operation)
                            )

                except ET.ParseError as err:
                    self.log(f"❌ XML parse error: {err}", "ERROR")

        except Exception as err:
            self.log(f"❌ Помилка читання файлу: {err}", "ERROR")

    # ══════════════════════════════════════════════════════════════════════════
    #  РЕНДЕР ТАБЛИЦІ
    # ══════════════════════════════════════════════════════════════════════════
    def _render_table(self):
        for w in self.rows_frame.winfo_children():
            w.destroy()

        df = pd.DataFrame(
            self.sales_data,
            columns=["Дата", "Час", "Номер чека", "Найменування", "Сума (грн)", "Тип операції"],
        ).sort_values(by=["Дата", "Час"])

        for date, group in df.groupby("Дата"):
            self._add_section_header(f"📅  {date}", COLORS["accent_blue"])

            for _, row in group.iterrows():
                rtype = "return" if row["Тип операції"] == "Повернення" else "normal"
                self._add_row(list(row), row_type=rtype)

            self._add_separator()

            totals = self.sales_totals_by_date.get(date, {})
            total_sales   = totals.get("Продаж", 0)
            total_returns = totals.get("Повернення", 0)
            balance       = total_sales - total_returns
            taxes         = totals.get("taxes", {})

            self._add_row(["", "", "", "Загальний обіг (Продаж)", f"{total_sales:.2f}", ""], "summary")

            for tax_char, td in sorted(taxes.items()):
                tv  = td.get("turnover", 0.0)
                vat = td.get("vat", 0.0)
                pr  = td.get("pr", "")
                if tv != 0:
                    self._add_row(["", "", "", f"  Обіг  Група {tax_char} ({pr})", f"{tv:.2f}", ""], "summary")
                if vat != 0:
                    self._add_row(["", "", "", f"  Податок Група {tax_char} ({pr})", f"{vat:.2f}", ""], "summary")

            self._add_row(["", "", "", "Повернення", f"{total_returns:.2f}", ""], "summary")
            self._add_row(["", "", "", "ЧИСТИЙ БАЛАНС", f"{balance:.2f}", ""], "summary")
            self._add_separator(COLORS["accent_blue"])

        # ── ЗВЕДЕНА ТАБЛИЦЯ ───────────────────────────────────────────────────
        self._add_section_header("▓▓  ЗВЕДЕНА ТАБЛИЦЯ ЗА ВЕСЬ ПЕРІОД  ▓▓", COLORS["accent_purple"])

        grand_sales   = sum(v.get("Продаж", 0)    for v in self.sales_totals_by_date.values())
        grand_returns = sum(v.get("Повернення", 0) for v in self.sales_totals_by_date.values())
        grand_taxes   = self._calc_grand_taxes()

        self._add_row(["", "", "", "ЗАГАЛЬНИЙ ПРОДАЖ", f"{grand_sales:.2f}", ""], "grand")

        for tax_char, td in sorted(grand_taxes.items()):
            tv  = td.get("turnover", 0.0)
            vat = td.get("vat", 0.0)
            pr  = td.get("pr", "")
            if tv != 0:
                self._add_row(["", "", "", f"  ОБІГ ГРУПА {tax_char} ({pr})", f"{tv:.2f}", ""], "grand")
            if vat != 0:
                self._add_row(["", "", "", f"  ПОДАТОК ГРУПА {tax_char} ({pr})", f"{vat:.2f}", ""], "grand")

        self._add_row(["", "", "", "ЗАГАЛЬНІ ПОВЕРНЕННЯ", f"{grand_returns:.2f}", ""], "grand")
        self._add_row(["", "", "", "ФІНАЛЬНИЙ БАЛАНС", f"{(grand_sales - grand_returns):.2f}", ""], "grand")

    # ══════════════════════════════════════════════════════════════════════════
    #  HELPER: ЗВЕДЕНІ ПОДАТКИ (завжди з актуальними ставками з _tax_rate_map)
    # ══════════════════════════════════════════════════════════════════════════
    def _calc_grand_taxes(self):
        """Підсумовує обороти і ПДВ по всіх днях.
        Ставку ЗАВЖДИ бере з self._tax_rate_map (не з збереженого pr),
        щоб уникнути "0.00%" якщо pr було записано до заповнення карти ставок."""
        tax_name_to_code = {v: k for k, v in TAX_MAP.items()}
        grand = {}
        for day_data in self.sales_totals_by_date.values():
            for tax_char, td in day_data.get("taxes", {}).items():
                code = tax_name_to_code.get(tax_char, "")
                pct  = self._tax_rate_map.get(code, 0.0)
                pr   = f"{pct:.2f}%"
                tv   = td.get("turnover", 0.0)
                vat  = tv * pct / (100 + pct) if pct > 0 else 0.0
                if tax_char not in grand:
                    grand[tax_char] = {"turnover": 0.0, "vat": 0.0, "pr": pr}
                grand[tax_char]["turnover"] += tv
                grand[tax_char]["vat"]      += vat
                grand[tax_char]["pr"]        = pr   # завжди актуальна ставка
        return grand

    # ══════════════════════════════════════════════════════════════════════════
    #  СТАТИСТИКА (ПРАВА ПАНЕЛЬ)
    # ══════════════════════════════════════════════════════════════════════════
    def _update_stats(self):
        total_sales   = sum(v.get("Продаж", 0)    for v in self.sales_totals_by_date.values())
        total_returns = sum(v.get("Повернення", 0) for v in self.sales_totals_by_date.values())
        net           = total_sales - total_returns
        checks        = len(set((r[0], r[2]) for r in self.sales_data))
        days          = len(self.sales_totals_by_date)

        self._stat_labels["total_checks"].configure( text=f"{checks}")
        self._stat_labels["total_sales"].configure(  text=f"{total_sales:,.2f} ₴")
        self._stat_labels["total_returns"].configure(text=f"{total_returns:,.2f} ₴")
        self._stat_labels["net_balance"].configure(  text=f"{net:,.2f} ₴")
        self._stat_labels["days_count"].configure(   text=f"{days}")

        # Очищаємо і заново будуємо блок податків
        for w in self.tax_stats_frame.winfo_children():
            w.destroy()

        grand_taxes = self._calc_grand_taxes()

        for tax_char, td in sorted(grand_taxes.items()):
            card = ctk.CTkFrame(self.tax_stats_frame, fg_color=COLORS["bg_card2"], corner_radius=8)
            card.pack(fill="x", padx=12, pady=3)

            ctk.CTkLabel(
                card,
                text=f"Група {tax_char}  ({td.get('pr', '')})",
                font=ctk.CTkFont(family="Consolas", size=10, weight="bold"),
                text_color=COLORS["accent_purple"],
                anchor="w",
            ).pack(fill="x", padx=10, pady=(6, 0))

            ctk.CTkLabel(
                card,
                text=f"Обіг:    {td['turnover']:,.2f} ₴",
                font=ctk.CTkFont(family="Consolas", size=10),
                text_color=COLORS["text_secondary"],
                anchor="w",
            ).pack(fill="x", padx=10)

            ctk.CTkLabel(
                card,
                text=f"Податок: {td['vat']:,.2f} ₴",
                font=ctk.CTkFont(family="Consolas", size=11, weight="bold"),
                text_color=COLORS["accent_yellow"],
                anchor="w",
            ).pack(fill="x", padx=10, pady=(0, 6))

    # ══════════════════════════════════════════════════════════════════════════
    #  ОЧИЩЕННЯ
    # ══════════════════════════════════════════════════════════════════════════
    def clear_data(self, silent=False):
        self.sales_data.clear()
        self.sales_totals_by_date.clear()
        self._tax_rate_map.clear()

        for w in self.rows_frame.winfo_children():
            w.destroy()

        for key, lbl in self._stat_labels.items():
            lbl.configure(text="—")

        for w in self.tax_stats_frame.winfo_children():
            w.destroy()

        self.progress.set(0)
        self.btn_export.configure(state="disabled")

        if not silent:
            self.log("🗑 Дані очищено.", "INFO")
            self.status_label.configure(text="Очікування файлу…")

    # ══════════════════════════════════════════════════════════════════════════
    #  ЕКСПОРТ В EXCEL
    # ══════════════════════════════════════════════════════════════════════════
    def export_to_excel(self):
        if not self.sales_data:
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel файли", "*.xlsx")],
        )
        if not save_path:
            return

        df = pd.DataFrame(
            self.sales_data,
            columns=["Дата", "Час", "Номер чека", "Найменування", "Сума (грн)", "Тип операції"],
        ).sort_values(by=["Дата", "Час"])

        grand_sales   = 0.0
        grand_returns = 0.0
        output_rows   = []

        for date, group in df.groupby("Дата"):
            output_rows.extend(group.values.tolist())

            totals        = self.sales_totals_by_date.get(date, {})
            total_sales   = totals.get("Продаж", 0)
            total_returns = totals.get("Повернення", 0)
            balance       = total_sales - total_returns
            taxes         = totals.get("taxes", {})

            grand_sales   += total_sales
            grand_returns += total_returns

            output_rows.append(["", "", "", f"--- ПІДСУМКИ ДНЯ {date} ---", "", ""])
            output_rows.append(["", "", "", "Загальний обіг (Продаж)", f"{total_sales:.2f}", ""])

            for tax_char, td in sorted(taxes.items()):
                tv  = td.get("turnover", 0.0)
                vat = td.get("vat", 0.0)
                pr  = td.get("pr", "")
                if tv != 0:
                    output_rows.append(["", "", "", f"Обіг Група {tax_char} ({pr})", f"{tv:.2f}", ""])
                if vat != 0:
                    output_rows.append(["", "", "", f"Податок Група {tax_char} ({pr})", f"{vat:.2f}", ""])

            output_rows.append(["", "", "", "Повернення", f"{total_returns:.2f}", ""])
            output_rows.append(["", "", "", "ЧИСТИЙ БАЛАНС", f"{balance:.2f}", ""])
            output_rows.append(["", "", "", "", "", ""])

        # Зведена таблиця — використовуємо _calc_grand_taxes для актуальних ставок
        grand_taxes = self._calc_grand_taxes()
        output_rows.append(["", "", "", "ЗВЕДЕНА ТАБЛИЦЯ ЗА ВЕСЬ ПЕРІОД", "", ""])
        output_rows.append(["", "", "", "ЗАГАЛЬНИЙ ПРОДАЖ", f"{grand_sales:.2f}", ""])

        for tax_char, td in sorted(grand_taxes.items()):
            tv  = td.get("turnover", 0.0)
            vat = td.get("vat", 0.0)
            pr  = td.get("pr", "")
            if tv != 0:
                output_rows.append(["", "", "", f"ЗАГАЛЬНИЙ ОБІГ ГРУПА {tax_char} ({pr})", f"{tv:.2f}", ""])
            if vat != 0:
                output_rows.append(["", "", "", f"ЗАГАЛЬНИЙ ПОДАТОК ГРУПА {tax_char} ({pr})", f"{vat:.2f}", ""])

        output_rows.append(["", "", "", "ЗАГАЛЬНІ ПОВЕРНЕННЯ",       f"{grand_returns:.2f}", ""])
        output_rows.append(["", "", "", "ФІНАЛЬНИЙ БАЛАНС", f"{(grand_sales - grand_returns):.2f}", ""])

        export_df = pd.DataFrame(
            output_rows,
            columns=["Дата", "Час", "Номер чека", "Найменування", "Сума (грн)", "Тип операції"],
        )

        try:
            export_df.to_excel(save_path, index=False)
        except PermissionError:
            messagebox.showerror("Помилка", "Файл відкритий. Закрийте Excel і спробуйте знову.")
            return

        # ── Форматування Excel ────────────────────────────────────────────────
        wb = load_workbook(save_path)
        ws = wb.active

        # Стилі заливок
        fills = {
            "return":  PatternFill("solid", fgColor="FFCCCC"),
            "summary": PatternFill("solid", fgColor="FFFACD"),
            "grand":   PatternFill("solid", fgColor="C6EFCE"),
            "header":  PatternFill("solid", fgColor="1F3864"),
        }
        thin = Side(style="thin", color="CCCCCC")
        border = Border(bottom=thin)

        # Заголовок
        for cell in ws[1]:
            cell.fill  = fills["header"]
            cell.font  = Font(bold=True, color="FFFFFF", name="Consolas")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            label = str(row[3].value) if row[3].value else ""
            op    = str(row[5].value) if row[5].value else ""

            if op == "Повернення":
                for cell in row:
                    cell.fill = fills["return"]
            elif "ПІДСУМКИ" in label or "Обіг" in label or "БАЛАНС" in label or "Податок" in label or "Повернення" in label:
                for cell in row:
                    cell.fill = fills["summary"]
                    cell.font = Font(bold="БАЛАНС" in label, name="Consolas", size=10)
            elif "ЗАГАЛЬНИЙ" in label or "ЗВЕДЕНА" in label or "ФІНАЛЬНИЙ" in label:
                for cell in row:
                    cell.fill = fills["grand"]
                    cell.font = Font(bold=True, name="Consolas", size=10)

        # Ширина колонок
        col_widths = [12, 10, 12, 50, 14, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(save_path)
        self.log(f"💾 Збережено: {save_path}", "OK")

        if messagebox.askyesno("Готово", "Файл збережено. Відкрити зараз?"):
            try:
                os.startfile(save_path)
            except Exception:
                messagebox.showwarning("Увага", "Не вдалося відкрити файл автоматично.")

        if self.temp_dir and os.path.exists(self.temp_dir):
            if messagebox.askyesno("Очищення", "Видалити тимчасові файли?"):
                try:
                    shutil.rmtree(self.temp_dir)
                    self.log("🧹 Тимчасові файли видалено.", "OK")
                    self.temp_dir = None
                except Exception as err:
                    messagebox.showwarning("Увага", f"Не вдалося видалити:\n{err}")


# ─── ЗАПУСК ───────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = SalesParserApp()
    app.mainloop()
