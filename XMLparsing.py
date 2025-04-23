import os
import re
import zipfile
import shutil
import tempfile
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import filedialog, messagebox
from ttkbootstrap import Style
from ttkbootstrap.widgets import Treeview, Button
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import webbrowser
import logging


def log_message(text_widget, message, level="INFO"):
    def open_url(url):
        webbrowser.open_new(url)

    text_widget.config(state="normal")
    if level == "INFO":
        logging.info(message)
    elif level == "WARNING":
        logging.warning(message)
    elif level == "ERROR":
        logging.error(message)

    if "http" in message:
        parts = message.split("http", 1)
        text_widget.insert("end", parts[0])
        url = "http" + parts[1]
        tag_name = f"link_{len(message)}"
        text_widget.insert("end", url + "\n", tag_name)
        text_widget.tag_config(tag_name, foreground="white", underline=True)
        text_widget.tag_bind(tag_name, "<Button-1>", lambda e, link=url: open_url(link))
    else:
        text_widget.insert("end", message + "\n")

    text_widget.config(state="disabled")
    text_widget.see("end")
    text_widget.update()

class SalesParserApp:
    def __init__(self, root):
        self.root = root
        self.root.title("XML Парсер Марія-304Т3")
        self.style = Style("superhero")
        self.sales_data = []
        self.temp_dir = None

        self.frame = tk.Frame(self.root)
        self.frame.pack(pady=10)

        self.select_btn = Button(self.frame, text="Обрати ZIP", bootstyle="warning", command=self.select_zip)
        self.select_btn.grid(row=0, column=0, padx=10)

        self.export_btn = Button(self.frame, text="Експорт в Excel", bootstyle="success", command=self.export_to_excel, state="disabled")
        self.export_btn.grid(row=0, column=1, padx=10)

        self.progress_var = tk.DoubleVar()
        self.progress = tk.ttk.Progressbar(self.root, variable=self.progress_var, maximum=100)
        self.progress.pack(fill="x", padx=10, pady=(0, 5))

        self.tree = Treeview(self.root, columns=("date", "time", "check", "name", "amount", "type"), show="headings", height=15)

        self.tree.heading("date", text="Дата")
        self.tree.column("date", width=90, anchor="center")

        self.tree.heading("time", text="Час")
        self.tree.column("time", width=80, anchor="center")

        self.tree.heading("check", text="Номер чека")
        self.tree.column("check", width=100, anchor="center")

        self.tree.heading("name", text="Найменування")
        self.tree.column("name", width=300, anchor="w", stretch=True)

        self.tree.heading("amount", text="Сума (грн)")
        self.tree.column("amount", width=90, anchor="e")

        self.tree.heading("type", text="Тип операції")
        self.tree.column("type", width=100, anchor="center")

        self.tree.pack(padx=10, pady=5, fill="both", expand=True)


        self.log_text = tk.Text(self.root, height=6, state="disabled", bg="black", fg="white", wrap="word")
        self.log_text.pack(fill="both", padx=10, pady=(0, 10))

    def select_zip(self):
        zip_path = filedialog.askopenfilename(filetypes=[("ZIP архів", "*.zip")])
        if not zip_path:
            return

        self.sales_data.clear()
        self.tree.delete(*self.tree.get_children())

        self.temp_dir = tempfile.mkdtemp()
        log_message(self.log_text, f"📦 Обрано архів: {zip_path}")

        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(self.temp_dir)
            log_message(self.log_text, f"✅ Розпаковано до: {self.temp_dir}")
        except zipfile.BadZipFile:
            messagebox.showerror("Помилка", "ZIP-файл пошкоджено або не підтримується.")
            return


        files = [f for f in os.listdir(self.temp_dir) if f.endswith(".xml")]
        total = len(files)
        if total == 0:
            log_message(self.log_text, "❌ XML-файли не знайдені у ZIP.")
            return

        log_message(self.log_text, f"🔍 Знайдено {total} XML-файлів. Починаємо обробку...")
        self.progress_var.set(0)
        self.progress.update()

        for idx, filename in enumerate(files, 1):
            file_path = os.path.join(self.temp_dir, filename)
            self.parse_file(file_path)
            self.progress_var.set((idx / total) * 100)
            self.progress.update()
            log_message(self.log_text, f"🗂 Оброблено: {filename}")

        for row in self.sales_data:
            self.tree.insert("", "end", values=row)

        if self.sales_data:
            self.export_btn.config(state="normal")
            log_message(self.log_text, "✅ Всі файли оброблено.")
        else:
            log_message(self.log_text, "⚠️ Файли прочитано, але не знайдено чеків.")

    def parse_file(self, filepath):
        try:
            with open(filepath, encoding="utf-8") as f:
                content = f.read()

            dat_blocks = re.findall(r'<DAT.*?</DAT>', content, re.DOTALL)

            for block in dat_blocks:
                fake_xml = f"<root>{block}</root>"
                try:
                    root = ET.fromstring(fake_xml)

                    for c_block in root.findall(".//C"):
                        items = c_block.findall(".//P")
                        e = c_block.find(".//E")
                        
                        if e is None:
                            continue  # если нет информации о чеке — пропускаем

                        ts = e.attrib.get("TS", "")
                        check_no = e.attrib.get("NO", "")
                        
                        if ts and len(ts) == 14:
                            date = f"{ts[:4]}-{ts[4:6]}-{ts[6:8]}"
                            time_str = f"{ts[8:10]}:{ts[10:12]}:{ts[12:]}"
                        else:
                            date = "Невідомо"
                            time_str = ""

                        for item in items:
                            name = item.attrib.get("NM", "Без назви")
                            amount_raw = int(item.attrib.get("SM", 0))
                            amount = abs(amount_raw) / 100
                            operation_type = "Продаж" if amount_raw >= 0 else "Повернення"

                            self.sales_data.append((date, time_str, check_no, name, f"{amount:.2f}", operation_type))

                except ET.ParseError as e:
                    log_message(self.log_text, f"❌ Помилка XML у {os.path.basename(filepath)}: {e}", level="ERROR")

        except Exception as e:
            log_message(self.log_text, f"❌ Помилка читання {os.path.basename(filepath)}: {e}", level="ERROR")

    def export_to_excel(self):
        if not self.sales_data:
            return

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel файли", "*.xlsx")])
        if not save_path:
            return

        df = pd.DataFrame(self.sales_data, columns=["Дата", "Час", "Номер чека", "Найменування", "Сума (грн)", "Тип операції"])
        df = df.sort_values(by=["Дата", "Час"])

        output_rows = []
        for date, group in df.groupby("Дата"):
            output_rows.extend(group.values.tolist())

            total_sales = group[group["Тип операції"] == "Продаж"]["Сума (грн)"].astype(float).sum()
            total_returns = group[group["Тип операції"] == "Повернення"]["Сума (грн)"].astype(float).sum()
            balance = total_sales - total_returns

            output_rows.append(["", "", "", "Ітого (продажі)", f"{total_sales:.2f}", ""])
            output_rows.append(["", "", "", "Ітого (повернення)", f"{total_returns:.2f}", ""])
            output_rows.append(["", "", "", "Баланс", f"{balance:.2f}", ""])
            output_rows.append(["", "", "", "", "", ""])

        export_df = pd.DataFrame(output_rows, columns=["Дата", "Час", "Номер чека", "Найменування", "Сума (грн)", "Тип операції"])

        try:
            export_df.to_excel(save_path, index=False)
        except PermissionError:
            messagebox.showerror("Помилка", "Не вдалося зберегти файл. Можливо, він відкритий у Excel.")
            return

        wb = load_workbook(save_path)
        ws = wb.active
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[5].value == "Повернення":
                for cell in row:
                    cell.fill = red_fill

        wb.save(save_path)
        log_message(self.log_text, f"💾 Excel-файл збережено: {save_path}")

        if messagebox.askyesno("Готово", "Файл збережено. Відкрити зараз?"):
            try:
                os.startfile(save_path)
            except Exception:
                messagebox.showwarning("Увага", "Не вдалося відкрити файл автоматично.")

        if self.temp_dir and os.path.exists(self.temp_dir):
            if messagebox.askyesno("Очищення", "Очистити тимчасові файли?"):
                try:
                    shutil.rmtree(self.temp_dir)
                    log_message(self.log_text, "🧹 Тимчасову папку очищено.")
                    self.temp_dir = None
                except Exception as e:
                    messagebox.showwarning("Увага", f"Не вдалося видалити тимчасові файли.\n{e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = SalesParserApp(root)
    root.mainloop()
